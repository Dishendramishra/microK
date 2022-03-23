from PySide2 import QtGui
from PySide2 import QtCore
from PySide2.QtWidgets import *
from PySide2.QtCore import *
from PySide2.QtGui import *

from qtpy import uic
import pyqtgraph as pg
from qt_material import apply_stylesheet

import serial
from openpyxl import Workbook

import sys
from time import sleep
from collections import deque
from datetime import date, datetime


if sys.platform == "linux" or sys.platform == "linux2":
    pass

elif sys.platform == "win32":
    import ctypes
    myappid = u'prl.microk'  # arbitrary string
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

elif sys.platform == "darwin":
    pass

class serialThread(QThread):
    progress_output_signal     = QtCore.Signal(list)
    graph_datapoint_signal     = QtCore.Signal(list)
    log_dataobject_signal      = QtCore.Signal(list)
    rw_channels_signal         = QtCore.Signal(bool)
    finished_signal            = QtCore.Signal(bool)

    def __init__(self):
        QThread.__init__(self)

        self.read_channels  = None
        self.write_channels = None
        self.log_channels   = None
        self.ser_read       = None
        self.ser_write      = None
        self.serial_flag    = False

    def __del__(self):
        print("Thread terminating...")

    def kill(self):
        try:
            self.ser_read.close()
            self.ser_write.close()
        except serial.serialutil.PortNotOpenError:
            print("Already Closed!")
        self.serial_flag  = False

    def initiate(self, readport, writeport):
        try:
            self.ser_read  = serial.Serial(readport, 9600)
            self.ser_write = serial.Serial(writeport, 9600)

        except:
            print("Invalid port!")
            msg = QMessageBox()
            msg.setWindowTitle("Error")
            msg.setText("No Device Found!")
            msg.setIcon(QMessageBox.Critical)
            msg.exec_()
            self.finished.emit(True)
            return None

        self.serial_flag = True

        # don't call run() directly, instead call start()
        # start() will autonatically run()
        self.start()

    def run(self):

        x_ = 1

        while self.serial_flag:
            self.rw_channels_signal.emit(True)
            if not self.read_channels:
                continue
            
            print("\nread  :",self.read_channels)
            print("write :",self.write_channels)
            print("log  :",self.log_channels,"\n")

            for channel_number in self.read_channels:
                try:
                    cmd = 'READ{}?\r\n'.format(channel_number)
                    self.ser_read.write(cmd.encode())
                    self.ser_read.flush()

                    output = self.ser_read.read_until(b'\r').decode().strip()
                    
                    if "not enabled" in output:
                        self.progress_output_signal.emit([channel_number, "Not Enabled!"])

                    else:
                        self.progress_output_signal.emit([channel_number, output])
                        
                        val = round(float(output.split(",")[0]),4)
                        numb_val =  val
                        val = str(val)+"\n"
                        print(f"CH-{channel_number} numb_val: {numb_val}")
                        
                        if channel_number in self.log_channels:
                            self.log_dataobject_signal.emit([channel_number, x_, numb_val])

                        # Plotting only write channels
                        if channel_number in self.write_channels:
                            self.graph_datapoint_signal.emit([ x_ , numb_val])
                            self.ser_write.write(val.encode())
                            self.ser_write.write(b"\n")
                            self.ser_write.flush()
                            x_ += 1                

                except Exception as e:
                            print(e)

class Ui(QMainWindow):
    
    def __init__(self, *args, **kwargs):
        super(Ui, self).__init__(*args, **kwargs)

        uic.loadUi('main_ui.ui', self)
        self.setWindowTitle("MickroK Temperature Reader")

        # ==========================================================
        #                       Graphing Window 
        # ==========================================================
        self.win = pg.GraphicsWindow()
        self.plot = self.win.addPlot(title='', labels={'bottom': 'time','left': "temperature"})
        self.curve = self.plot.plot()
        self.data = deque(maxlen=20)
        # ==========================================================


        # ==========================================================
        #                       Global Vars
        # ==========================================================
        self.val_lbls = {
            1   :   self.lbl_ch1_val,
            2   :   self.lbl_ch2_val,
            3   :   self.lbl_ch3_val,
            10  :   self.lbl_ch4_val,
            11  :   self.lbl_ch5_val,
            12  :   self.lbl_ch6_val,
            13  :   self.lbl_ch7_val,
            14  :   self.lbl_ch8_val,
            15  :   self.lbl_ch9_val,
            16  :   self.lbl_ch10_val
        }
        
        self.read_chbx_grp = {
            1   :   self.chbx_read1,
            2   :   self.chbx_read2,
            3   :   self.chbx_read3,
            10  :   self.chbx_read4,
            11  :   self.chbx_read5,
            12  :   self.chbx_read6,
            13  :   self.chbx_read7,
            14  :   self.chbx_read8,
            15  :   self.chbx_read9,
            16  :   self.chbx_read10,
        }

        self.write_chbx_grp = {
            1   :   self.chbx_write1,
            2   :   self.chbx_write2,
            3   :   self.chbx_write3,
            10  :   self.chbx_write4,
            11  :   self.chbx_write5,
            12  :   self.chbx_write6,
            13  :   self.chbx_write7,
            14  :   self.chbx_write8,
            15  :   self.chbx_write9,
            16  :   self.chbx_write10,
        }

        self.log_chbx_grp = {
            1   :   self.chbx_log1,
            2   :   self.chbx_log2,
            3   :   self.chbx_log3,
            10  :   self.chbx_log4,
            11  :   self.chbx_log5,
            12  :   self.chbx_log6,
            13  :   self.chbx_log7,
            14  :   self.chbx_log8,
            15  :   self.chbx_log9,
            16  :   self.chbx_log10,
        }

        # This variable is used to keep track of date while 
        # saving data in update_graph()
        self.workbook_create_date = None

        self.channels_to_read   = None 
        self.channels_to_write  = None 
        self.channels_to_log    = None

        self.channel_indices = None

        self.col_in_sheet = 1
        # ==========================================================
        #                 Thread Signals & Slots
        # ==========================================================
        self.serial_thread = serialThread()
        self.serial_thread.finished_signal.connect(self.stop)
        self.serial_thread.rw_channels_signal.connect(self.get_rw_channels)
        self.serial_thread.progress_output_signal.connect(self.update_channel_val)
        self.serial_thread.graph_datapoint_signal.connect(self.update_graph)
        self.serial_thread.log_dataobject_signal.connect(self.update_workbook)
        
        # ==========================================================

        self.btn_stop.setEnabled(False)


        # ==========================================================
        #                        GUI Slots
        # ==========================================================
        self.btn_start.pressed.connect(self.start)
        self.btn_stop.pressed.connect(self.stop)
        self.btn_clear.pressed.connect(self.clear)
        # ==========================================================

        self.show()

    def closeEvent(self,event):
        self.serial_thread.kill()
        while self.serial_thread.isRunning():
            sleep(0.1)

        self.win.close()

    def get_rw_channels(self):
        self.channels_to_read = []
        self.channels_to_write = []
        self.channels_to_log = []

        for key, value in self.read_chbx_grp.items():
                if value.isChecked():
                    self.channels_to_read.append(key)

        for key, value in self.write_chbx_grp.items():
                if value.isChecked():
                    self.channels_to_write.append(key)

        for key, value in self.log_chbx_grp.items():
            if value.isChecked():
                self.channels_to_log.append(key)

        self.serial_thread.read_channels  = self.channels_to_read
        self.serial_thread.write_channels = self.channels_to_write
        self.serial_thread.log_channels   = self.channels_to_log

        self.configure_workbook()

    def update_channel_val(self, data_list):
        self.val_lbls[data_list[0]].setText(data_list[1])


    def update_graph(self, data_point):
        print("update_graph(): ", data_point)
        val_x, val_y = data_point
        self.data.append({"x" : val_x, "y" : val_y})
        x = [item['x'] for item in self.data]
        y = [item['y'] for item in self.data]
        self.curve.setData(x,y)


    # ==========================================================
    #                 Excel Workbook Functions
    # ==========================================================    
    def create_workbook(self):
        self.workbook = Workbook()
        self.workbook_create_date = datetime.now()
        self.sheet = self.workbook.active
        self.col_in_sheet = 1

    def configure_workbook(self):

        current_row = 2

        if not self.channel_indices:
            self.channel_indices = {}

        for channel in self.channels_to_log:

            # before adding a channel in excel file we need to find 
            # the starting row number for newly added column, to
            # start saving data
            if channel in self.channel_indices:
                current_row = self.channel_indices[channel][2]

            if channel not in self.channel_indices:
                #                               [ timestamp col, value col, row]
                self.sheet.cell(row=1, column=self.col_in_sheet).value = f'CH-{channel} Timestamp'
                self.sheet.cell(row=1, column=self.col_in_sheet+1).value =f'CH-{channel}'

                self.channel_indices[channel] = [self.col_in_sheet, self.col_in_sheet+1, current_row]
                
                self.col_in_sheet  += 2

    def update_workbook(self, data_obj):
        # Checking if new day starting, if new day start then 
        # creating and configuring a new excel file.
        if datetime.now().strftime("%d") > self.workbook_create_date.strftime("%d"):
            self.save_workbook()
            self.create_workbook()

        channel, val_x, val_y = data_obj
        timestamp = self.channel_indices[channel][0]
        value     = self.channel_indices[channel][1]
        row       = self.channel_indices[channel][2]

        ts = datetime.now().strftime("%Y-%m-%d:%H:%M:%S")
        print(f'CH-{channel} : {ts}')

        self.sheet.cell(row=row, column=timestamp).value = ts
        self.sheet.cell(row=row, column=value).value     = val_y

        # increading row value for writing next time
        self.channel_indices[channel][2] += 1

    def save_workbook(self):
        filename = self.workbook_create_date.strftime("%Y-%m-%d-%f")
        self.workbook.save(filename="{}.xlsx".format(filename))
        self.workbook.close()

    # ==========================================================


    # ==========================================================
    #                        GUI Events
    # ==========================================================    
    def start(self):
        self.data.clear()
        self.curve.clear()
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.btn_clear.setEnabled(False)
        read_port = self.cmb_readport.currentText()
        write_port = self.cmb_writeport.currentText()

        self.create_workbook()
        self.serial_thread.initiate(read_port, write_port)

    
    def stop(self):
        self.save_workbook()
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.btn_clear.setEnabled(True)
        self.serial_thread.kill()
    
    def clear(self):
        for ch_num in self.val_lbls:
            self.val_lbls[ch_num].setText("-")
    # ==========================================================

app = QApplication(sys.argv)
app.setWindowIcon(QIcon("resources/icons/prl.png"))
apply_stylesheet(app, theme='dark_blue.xml')
window = Ui()
app.exec_()
